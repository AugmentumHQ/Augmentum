"""Spreadsheet artifact tool — generates XLSX files from structured data."""

from __future__ import annotations

import io
import json
import re
from typing import TYPE_CHECKING

from augmentum.tools.base import Tool, ToolCategory, ToolResult
from augmentum.utils.logging import get_logger

if TYPE_CHECKING:
    from augmentum.tools.artifact_storage import ArtifactStore

log = get_logger(__name__)


class SpreadsheetTool(Tool):
    """Generate Excel spreadsheets from structured sheet definitions."""

    def __init__(self, artifact_store: ArtifactStore) -> None:
        self._store = artifact_store

    @property
    def name(self) -> str:
        return "create_spreadsheet"

    @property
    def description(self) -> str:
        return (
            "Create an Excel spreadsheet (.xlsx) with one or more sheets. "
            "Call it when the user wants data they can keep, sort, filter, or "
            "open elsewhere — an export, a tracker, a budget, a comparison "
            "table they'll work in. For data meant only to be READ in the "
            "reply, write a markdown table instead; for data meant to be SEEN "
            "as a shape or trend, call create_chart. "
            "Each sheet has a name, column headers, and rows of data. "
            "Supports styling (bold headers, auto-width columns), column number "
            "formats, Excel formulas (cells starting with '='), and summary rows "
            "(sum/average/count). Returns a download link for the generated file."
            " Use actual data values — no placeholder text like TBD, [Insert], or N/A."
        )

    @property
    def category(self) -> ToolCategory:
        return ToolCategory.ARTIFACT

    @property
    def error_hints(self) -> dict[str, str]:
        """Recovery guidance for the errors this tool actually emits.

        Same contract as ChartTool.error_hints — keys must remain substrings of
        real error strings so ``Tool.enrich_error`` can match them.
        """
        return {
            "No sheets provided": (
                "`sheets` must be an array of objects, each with `name`, "
                '`headers` (array of column names) and `rows` (array of arrays, '
                'one inner array per row): [{"name":"Q1","headers":["Item",'
                '"Cost"],"rows":[["Rent",1200]]}]. A single flat array of '
                "objects is not accepted — wrap it in a sheet."
            ),
            "No data after cleanup": (
                "Every row was empty or whitespace. Each row must be an array "
                "of cell values aligned to `headers`, and there must be at "
                "least one row with real data — placeholders like TBD or "
                "[Insert] are stripped."
            ),
            "No module named 'openpyxl'": (
                "The spreadsheet writer is not installed on this server, so no "
                "retry will succeed. Do NOT call this tool again in this "
                "conversation — present the data as a markdown table instead "
                "and tell the user .xlsx export is unavailable."
            ),
        }

    @property
    def model_hint(self) -> str:
        return (
            "Use actual values, not placeholders. Prefer ONE call with several "
            "entries in `sheets` over several calls. Do NOT reach for this just "
            "to show a few rows in conversation — that's a markdown table. It "
            "produces a FILE the user downloads, so it's the right choice only "
            "when they want to keep or edit the data."
        )

    def health_check(self) -> bool:
        """False when openpyxl isn't installed — same contract as ChartTool.

        openpyxl is lazy-imported at render time, so without this the tool
        would appear on the model's roster and fail at execute.
        """
        from importlib.util import find_spec
        try:
            return find_spec("openpyxl") is not None
        except (ImportError, ValueError):
            return False

    @property
    def input_schema(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "title": {
                    "type": "string",
                    "description": "Spreadsheet filename (without extension)",
                },
                "theme": {
                    "type": "string",
                    "description": "Visual theme preset (slate, corporate, modern, emerald, rose). Default: slate",
                    "default": "",
                },
                "sheets": {
                    "type": "array",
                    "description": "List of sheets to create",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {
                                "type": "string",
                                "description": "Sheet tab name",
                            },
                            "headers": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": "Column header names",
                            },
                            "rows": {
                                "type": "array",
                                "description": "Data rows (each row is an array of values)",
                                "items": {
                                    "type": "array",
                                    "items": {},
                                },
                            },
                            "freeze_header": {
                                "type": "boolean",
                                "description": "Freeze the header row (default: true)",
                                "default": True,
                            },
                            "column_formats": {
                                "type": "object",
                                "description": (
                                    "Column number formats keyed by header name "
                                    '(e.g. {"Revenue": "$#,##0.00", "Growth": "0.0%"})'
                                ),
                                "default": {},
                            },
                            "summary_row": {
                                "type": "string",
                                "enum": ["sum", "average", "count", "none"],
                                "description": "Add a summary row with the specified function (default: none)",
                                "default": "none",
                            },
                        },
                        "required": ["name", "headers", "rows"],
                    },
                },
            },
            "required": ["title", "sheets"],
        }

    async def execute(
        self,
        *,
        title: str = "Spreadsheet",
        theme: str = "",
        sheets: list | None = None,
        task_id: str = "",
        session_id: str = "",
        **kwargs,
    ) -> ToolResult:
        from augmentum.tools.artifact_normalize import normalize_sheets, normalize_str

        title = normalize_str(title, "Spreadsheet")
        sheets = normalize_sheets(sheets)
        if not sheets:
            return ToolResult(success=False, error="No sheets provided")

        from augmentum.tools.artifact_sanitize import sanitize_sheets
        sheets = sanitize_sheets(sheets)
        if not sheets:
            return ToolResult(success=False, error="No data after cleanup")

        # Honest low-data signal — covers both the direct tool-call path and the
        # build pipeline. We still render (the sheet keeps its headers) but flag
        # empty/thin data instead of claiming a header-only sheet is "ready".
        from augmentum.tools.artifact_validate import sheet_quality
        quality = sheet_quality(sheets)
        warnings: list[str] = []
        if quality.degenerate:
            warnings.append(
                f"Spreadsheet has limited data ({quality.reason}) — it may look "
                "sparse or empty. Try a more specific request or supply the rows "
                "you want filled in."
            )

        try:
            data = _render_xlsx(sheets, theme_name=theme)

            safe_title = re.sub(r'[^\w\s-]', '', title).strip().replace(' ', '_')[:60]
            filename = f"{safe_title}.xlsx"

            total_rows = sum(len(s.get("rows", [])) for s in sheets)

            info = await self._store.save(
                data=data,
                filename=filename,
                fmt="xlsx",
                task_id=task_id,
                session_id=session_id,
                display_name=f"{title}.xlsx",
                user_id=Tool.extract_user_id(kwargs),
                metadata={
                    "page_type": "spreadsheet",
                    "sheet_count": len(sheets),
                    "total_rows": total_rows,
                },
                source_json=json.dumps({
                    "type": "spreadsheet",
                    "title": title,
                    "theme": theme,
                    "sheets": sheets,
                }),
            )

            from augmentum.tools.base import (
                format_output_with_warnings,
                make_artifact_card,
            )

            if quality.degenerate:
                summary = (
                    f"Spreadsheet '{title}' was created, but it has limited data "
                    f"({quality.reason}) so it may look sparse or empty."
                )
            else:
                summary = (
                    f"Spreadsheet '{title}' is ready — "
                    f"{len(sheets)} sheet{'s' if len(sheets) != 1 else ''}, "
                    f"{total_rows} data row{'s' if total_rows != 1 else ''}. "
                    "Available in the artifact library."
                )
            card = make_artifact_card(
                info,
                kind="artifact",
                title=title,
                subtitle=f"{len(sheets)} sheet{'s' if len(sheets) != 1 else ''}",
                summary=summary,
                preview={
                    "artifact_kind": "spreadsheet",
                    "format": "xlsx",
                    "size_bytes": info.get("size_bytes", 0),
                    "sheets": [
                        {
                            "name": s.get("name", f"Sheet{i+1}"),
                            "headers": s.get("headers", [])[:8],
                            "row_count": len(s.get("rows", [])),
                        }
                        for i, s in enumerate(sheets)
                    ],
                    "total_rows": total_rows,
                    "low_data": quality.degenerate,
                    "quality_note": quality.reason,
                },
            )
            return ToolResult(
                success=True,
                output=format_output_with_warnings(summary, warnings),
                metadata=info,
                card=card,
                warnings=warnings,
            )
        except Exception as e:
            log.error("spreadsheet_creation_failed", error=str(e), exc_info=True)
            return ToolResult(success=False, error=f"Spreadsheet creation failed: {e}")


# ---------------------------------------------------------------------------
# XLSX rendering via openpyxl
# ---------------------------------------------------------------------------

# Excel/CSV injection denylist: function names that, if present in a formula,
# can exfiltrate data, fetch remote content, hijack cells, or invoke legacy
# COM/macro execution paths. Match is case-insensitive against the formula
# token stream — see _neutralize_cell_value below.
_DANGEROUS_FORMULA_FUNCS = (
    # Remote fetch / exfiltration
    "IMPORTXML", "IMPORTHTML", "IMPORTDATA", "IMPORTRANGE", "IMPORTFEED",
    "WEBSERVICE", "FILTERXML", "RTD", "ENCODEURL",
    # Cell-reference hijack used in CSV-injection chains
    "INDIRECT", "HYPERLINK",
    # Legacy macro / COM / shell vectors
    "CALL", "REGISTER.ID", "EXECUTE", "EVALUATE", "EXEC", "SHELL",
    "GETDATA", "DDE", "DDEAUTO",
)
_DANGEROUS_FUNC_RE = re.compile(
    r"\b(" + "|".join(re.escape(f) for f in _DANGEROUS_FORMULA_FUNCS) + r")\s*\(",
    re.IGNORECASE,
)


def _neutralize_cell_value(value):
    """Defuse CSV/Excel-injection payloads in user-supplied cell values.

    Excel treats any cell beginning with =, +, -, or @ as a formula. The
    OWASP-recommended mitigation is to prefix such values with a single
    quote so Excel renders them as literal text.

    We allow plausibly-legitimate arithmetic formulas (e.g. =SUM(A1:A10),
    =A1+B1) but reject formulas containing dangerous functions (IMPORTXML,
    INDIRECT, WEBSERVICE, etc.) or DDE-style pipe references.

    Returns (sanitized_value, was_blocked: bool).
    """
    if not isinstance(value, str) or not value:
        return value, False

    first = value[0]
    if first not in ("=", "+", "-", "@"):
        return value, False

    if first == "=":
        # Legitimate formula path — but block known-dangerous calls,
        # DDE-pipe payloads (e.g. =cmd|'/c calc'!A1), and external
        # workbook/DDE references (e.g. ='C:\\Windows\\cmd.exe'!A1).
        # The 'X'! pattern is Excel's external-reference syntax, which
        # an LLM has no legitimate reason to emit.
        if (
            "|" in value
            or "'!" in value
            or _DANGEROUS_FUNC_RE.search(value)
        ):
            log.warning(
                "spreadsheet_dangerous_formula_blocked",
                preview=value[:80],
            )
            return "'" + value, True
        return value, False

    # +/-/@ prefixes: keep negative numbers, neutralize everything else.
    try:
        float(value)
        return value, False
    except ValueError:
        return "'" + value, False


def _render_xlsx(sheets: list, theme_name: str = "", design: dict | None = None) -> bytes:
    """Render sheets to an XLSX file using openpyxl + theme system.

    `design` (optional) chooses the font face and size. line_height is a
    no-op for spreadsheets; density scales row height when present.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from augmentum.tools.artifact_theme import (
        FONT_FAMILY_XLSX,
        apply_design,
        get_theme,
    )

    theme = apply_design(get_theme(theme_name), design)
    family_pref = (design or {}).get("font_family", "system")
    base_font_name = FONT_FAMILY_XLSX.get(family_pref, "Calibri")
    # body_size baseline 10 → Calibri size 11 (Excel's standard) at scale 1.0
    base_font_size = max(8, int(round(theme.body_size + 1)))
    _header_fill_hex = theme.hex_no_hash(theme.xlsx_header_fill)
    _zebra_fill_hex = theme.hex_no_hash(theme.xlsx_zebra_fill)
    _border_hex = theme.hex_no_hash(theme.border)

    # density: Excel auto-sizes row height to match font size. font_size_scale
    # already passes through `theme.body_size`, so density.spacious/compact
    # tweaks here would double-count. Intentionally no explicit row height
    # override — keep XLSX behavior driven by the font size alone.

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(
        name=base_font_name, size=base_font_size, bold=True, color="FFFFFF",
    )
    header_fill = PatternFill(
        start_color=_header_fill_hex,
        end_color=_header_fill_hex,
        fill_type="solid",
    )
    zebra_fill = PatternFill(
        start_color=_zebra_fill_hex,
        end_color=_zebra_fill_hex,
        fill_type="solid",
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color=_border_hex),
    )
    summary_border = Border(
        top=Side(style="medium", color=theme.hex_no_hash(theme.text)),
        bottom=Side(style="thin", color=_border_hex),
    )
    data_font = Font(name=base_font_name, size=base_font_size)
    summary_font = Font(name=base_font_name, size=base_font_size, bold=True)

    for sheet_def in sheets:
        name = sheet_def.get("name", "Sheet")[:31]  # Excel 31-char limit
        headers = sheet_def.get("headers", [])
        rows = sheet_def.get("rows", [])
        freeze = sheet_def.get("freeze_header", True)
        column_formats = sheet_def.get("column_formats") or {}
        summary_row = sheet_def.get("summary_row", "none") or "none"

        ws = wb.create_sheet(title=name)

        # Build a mapping from header name to column index (1-based)
        header_col_map: dict[str, int] = {}
        for col_idx, header in enumerate(headers, 1):
            header_col_map[header] = col_idx

        # Build column index to number format mapping
        col_format_map: dict[int, str] = {}
        for hdr_name, fmt_str in column_formats.items():
            if hdr_name in header_col_map:
                col_format_map[header_col_map[hdr_name]] = fmt_str

        # Track max column widths for auto-sizing
        col_widths: dict[int, int] = {}

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            header, _ = _neutralize_cell_value(header)
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            col_widths[col_idx] = len(str(header)) + 2

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                value, _ = _neutralize_cell_value(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                # Zebra stripe: alternate row background for readability
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                # Apply number format if configured for this column
                if col_idx in col_format_map:
                    cell.number_format = col_format_map[col_idx]
                # Track width
                val_len = len(str(value)) + 2 if value is not None else 2
                if col_idx not in col_widths or val_len > col_widths[col_idx]:
                    col_widths[col_idx] = val_len

        # Summary row
        if summary_row != "none" and rows and headers:
            func_name = {"sum": "SUM", "average": "AVERAGE", "count": "COUNT"}[
                summary_row
            ]
            summary_row_idx = len(rows) + 2  # header=1, data starts at 2
            data_start_row = 2
            data_end_row = len(rows) + 1

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=summary_row_idx, column=col_idx)
                cell.font = summary_font
                cell.border = summary_border

                # Check if this column has numeric data worth summarizing
                col_letter = get_column_letter(col_idx)
                has_numeric = any(
                    isinstance(row[col_idx - 1], int | float)
                    for row in rows
                    if col_idx - 1 < len(row)
                )
                if has_numeric:
                    cell.value = (
                        f"={func_name}({col_letter}{data_start_row}"
                        f":{col_letter}{data_end_row})"
                    )
                    # Apply matching number format if configured
                    if col_idx in col_format_map:
                        cell.number_format = col_format_map[col_idx]
                elif col_idx == 1:
                    # Label the summary row in the first column
                    cell.value = func_name.capitalize()

        # Auto-size columns (cap at 50 chars)
        for col_idx, width in col_widths.items():
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(width, 50)

        # Freeze header row
        if freeze and headers:
            ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
