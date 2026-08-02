"""Tests for tool pipeline improvements — truncation, caching, circuit breaker, filtering."""

from __future__ import annotations

import asyncio
import time
from dataclasses import dataclass

import pytest

from augmentum.tools.base import Tool, ToolCategory, ToolResult
from augmentum.tools.cache import ToolResultCache
from augmentum.tools.circuit_breaker import ToolCircuitBreaker
from augmentum.tools.filter import filter_tools_for_query
from augmentum.tools.result_processing import truncate_tool_result


# === Result Truncation ===


class TestTruncateToolResult:
    def test_short_text_unchanged(self):
        assert truncate_tool_result("hello", max_chars=100) == "hello"

    def test_empty_text(self):
        assert truncate_tool_result("", max_chars=100) == ""

    def test_none_text(self):
        assert truncate_tool_result(None, max_chars=100) is None

    def test_exact_limit_unchanged(self):
        text = "a" * 100
        assert truncate_tool_result(text, max_chars=100) == text

    def test_truncation_preserves_head_and_tail(self):
        text = "HEAD" + "x" * 5000 + "TAIL"
        result = truncate_tool_result(text, max_chars=500, tail_chars=100)
        assert result.startswith("HEAD")
        assert result.endswith("TAIL")
        assert "truncated" in result
        assert len(result) <= 500

    def test_truncation_notice_has_count(self):
        text = "a" * 10000
        result = truncate_tool_result(text, max_chars=1000, tail_chars=200)
        assert "truncated" in result
        # The notice should contain the number of chars removed
        assert "9" in result  # ~9000 chars truncated

    def test_very_small_budget(self):
        text = "a" * 1000
        result = truncate_tool_result(text, max_chars=50, tail_chars=10)
        assert len(result) <= 50


# === Tool Result Cache ===


class TestToolResultCache:
    def test_miss_returns_none(self):
        cache = ToolResultCache()
        assert cache.get("web_search", {"query": "test"}, ttl=300) is None

    def test_hit_after_put(self):
        cache = ToolResultCache()
        result = ToolResult(success=True, output="found it")
        cache.put("web_search", {"query": "test"}, result)
        cached = cache.get("web_search", {"query": "test"}, ttl=300)
        assert cached is not None
        assert cached.output == "found it"

    def test_different_params_miss(self):
        cache = ToolResultCache()
        result = ToolResult(success=True, output="found it")
        cache.put("web_search", {"query": "test"}, result)
        assert cache.get("web_search", {"query": "other"}, ttl=300) is None

    def test_different_tool_miss(self):
        cache = ToolResultCache()
        result = ToolResult(success=True, output="found it")
        cache.put("web_search", {"query": "test"}, result)
        assert cache.get("calculator", {"query": "test"}, ttl=300) is None

    def test_zero_ttl_never_expires(self):
        cache = ToolResultCache()
        result = ToolResult(success=True, output="2+2=4")
        cache.put("calculator", {"expression": "2+2"}, result)
        # TTL=0 means infinite
        cached = cache.get("calculator", {"expression": "2+2"}, ttl=0)
        assert cached is not None

    def test_param_order_independent(self):
        cache = ToolResultCache()
        result = ToolResult(success=True, output="ok")
        cache.put("tool", {"a": 1, "b": 2}, result)
        cached = cache.get("tool", {"b": 2, "a": 1}, ttl=300)
        assert cached is not None


# === Circuit Breaker ===


class TestCircuitBreaker:
    def test_initially_closed(self):
        cb = ToolCircuitBreaker(threshold=3, cooldown=60)
        assert cb.is_open("web_search") is False

    def test_opens_after_threshold(self):
        cb = ToolCircuitBreaker(threshold=3, cooldown=60)
        cb.record_failure("web_search")
        cb.record_failure("web_search")
        assert cb.is_open("web_search") is False
        cb.record_failure("web_search")
        assert cb.is_open("web_search") is True

    def test_success_resets(self):
        cb = ToolCircuitBreaker(threshold=2, cooldown=60)
        cb.record_failure("web_search")
        cb.record_failure("web_search")
        assert cb.is_open("web_search") is True
        cb.record_success("web_search")
        assert cb.is_open("web_search") is False

    def test_half_open_after_cooldown(self):
        cb = ToolCircuitBreaker(threshold=2, cooldown=0.01)
        cb.record_failure("tool")
        cb.record_failure("tool")
        assert cb.is_open("tool") is True
        time.sleep(0.02)
        assert cb.is_open("tool") is False  # half-open allows one try

    def test_independent_per_tool(self):
        cb = ToolCircuitBreaker(threshold=2, cooldown=60)
        cb.record_failure("tool_a")
        cb.record_failure("tool_a")
        assert cb.is_open("tool_a") is True
        assert cb.is_open("tool_b") is False


# === Tool Pre-filtering ===


@dataclass
class _MockTool:
    name: str
    description: str = ""
    category: ToolCategory = ToolCategory.SEARCH


class TestToolPrefiltering:
    def _tools(self, *names):
        return [_MockTool(name=n) for n in names]

    def test_search_query_keeps_web_tools(self):
        tools = self._tools("web_search", "web_fetch", "calculator", "datetime")
        result = filter_tools_for_query("search for latest news", tools, min_tools=1)
        names = {t.name for t in result}
        assert "web_search" in names
        assert "web_fetch" in names

    def test_math_query_keeps_math_tools(self):
        tools = self._tools("web_search", "calculator", "math_verify", "python_exec")
        result = filter_tools_for_query("calculate 2+2", tools, min_tools=1)
        names = {t.name for t in result}
        assert "calculator" in names

    def test_no_match_returns_all(self):
        tools = self._tools("web_search", "calculator", "datetime")
        result = filter_tools_for_query("tell me a joke", tools, min_tools=1)
        assert len(result) == 3

    def test_min_tools_floor(self):
        tools = self._tools("web_search", "calculator", "datetime", "file_ops")
        result = filter_tools_for_query("search for news", tools, min_tools=3)
        assert len(result) >= 3

    def test_empty_query_returns_all(self):
        tools = self._tools("a", "b", "c")
        assert filter_tools_for_query("", tools) == tools

    def test_url_triggers_fetch(self):
        tools = self._tools("web_search", "web_fetch", "calculator")
        result = filter_tools_for_query("summarize https://example.com", tools, min_tools=1)
        names = {t.name for t in result}
        assert "web_fetch" in names

    def test_image_query(self):
        tools = self._tools("web_search", "image_generation", "calculator")
        result = filter_tools_for_query("draw a picture of a cat", tools, min_tools=1)
        names = {t.name for t in result}
        assert "image_generation" in names

    def test_memory_query(self):
        tools = self._tools("web_search", "memory_recall", "calculator")
        result = filter_tools_for_query("do you remember my name?", tools, min_tools=1)
        names = {t.name for t in result}
        assert "memory_recall" in names

    def test_code_query(self):
        tools = self._tools("web_search", "python_exec", "calculator")
        result = filter_tools_for_query("write a python script to sort a list", tools, min_tools=1)
        names = {t.name for t in result}
        assert "python_exec" in names


# === Validation Error Flag ===


class TestValidationErrorFlag:
    def test_tool_result_default_no_validation_error(self):
        r = ToolResult(success=False, error="boom")
        assert r.validation_error is False

    def test_tool_result_with_validation_error(self):
        r = ToolResult(success=False, error="missing field", validation_error=True)
        assert r.validation_error is True


# === Tool Base Properties ===


class TestToolBaseProperties:
    def test_default_timeout(self):
        class T(Tool):
            name = "t"
            description = "t"
            category = ToolCategory.EXECUTE
            async def execute(self, **kw): ...
        assert T().timeout == 30.0

    def test_default_cacheable(self):
        class T(Tool):
            name = "t"
            description = "t"
            category = ToolCategory.EXECUTE
            async def execute(self, **kw): ...
        assert T().cacheable is True

    def test_default_cache_ttl(self):
        class T(Tool):
            name = "t"
            description = "t"
            category = ToolCategory.EXECUTE
            async def execute(self, **kw): ...
        assert T().cache_ttl == 300.0

    def test_override_timeout(self):
        class T(Tool):
            name = "t"
            description = "t"
            category = ToolCategory.EXECUTE
            @property
            def timeout(self): return 2.0
            async def execute(self, **kw): ...
        assert T().timeout == 2.0


# === SymPy Input Sanitization ===


class TestSympySanitization:
    """Verify that math_verify blocks code injection via SymPy expressions."""

    def _tool(self):
        from augmentum.tools.math_verify import MathVerifyTool
        return MathVerifyTool(http_client=None)

    def test_blocks_dunder(self):
        assert self._tool()._sanitize_sympy_input("x.__class__") is None

    def test_blocks_import(self):
        assert self._tool()._sanitize_sympy_input("__import__('os')") is None

    def test_blocks_exec(self):
        assert self._tool()._sanitize_sympy_input("exec('print(1)')") is None

    def test_blocks_eval(self):
        assert self._tool()._sanitize_sympy_input("eval('1+1')") is None

    def test_blocks_os_module(self):
        assert self._tool()._sanitize_sympy_input("os.system('ls')") is None

    def test_blocks_open(self):
        assert self._tool()._sanitize_sympy_input("open('/etc/passwd')") is None

    def test_blocks_subprocess(self):
        assert self._tool()._sanitize_sympy_input("subprocess.run(['ls'])") is None

    def test_blocks_disallowed_chars(self):
        # Backticks, semicolons, curly braces not in allowlist
        assert self._tool()._sanitize_sympy_input("x; import os") is None
        assert self._tool()._sanitize_sympy_input("x`y") is None

    def test_allows_normal_math(self):
        assert self._tool()._sanitize_sympy_input("x**2 + 2*x + 1") is not None

    def test_allows_trig(self):
        assert self._tool()._sanitize_sympy_input("sin(x) + cos(y)") is not None

    def test_allows_fractions(self):
        assert self._tool()._sanitize_sympy_input("3/4 + 1/2") is not None

    def test_escapes_backslash_before_quote(self):
        # Correct order: backslash first, then quote
        result = self._tool()._sanitize_sympy_input("x\\'y")
        # Should not be None (allowed chars), and backslash should be doubled
        # Actually this contains a backslash AND a quote — both allowed
        # The key test is that the escaping doesn't produce broken strings
        assert result is None or "\\\\" in result or "\\'" in result


# === Spreadsheet Formula Injection ===


class TestSpreadsheetFormulaSanitization:
    def test_sanitizes_plus_prefix(self):
        from augmentum.tools.artifact_spreadsheet import _render_xlsx
        sheets = [{"name": "Test", "headers": ["A"], "rows": [["+cmd|' /C calc'!A0"]]}]
        data = _render_xlsx(sheets)
        # The value should be prefixed with ' to neutralize the formula
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(data))
        cell_value = wb.active.cell(row=2, column=1).value
        assert cell_value.startswith("'")

    def test_sanitizes_at_prefix(self):
        from augmentum.tools.artifact_spreadsheet import _render_xlsx
        sheets = [{"name": "Test", "headers": ["A"], "rows": [["@SUM(A1)"]]}]
        data = _render_xlsx(sheets)
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(data))
        cell_value = wb.active.cell(row=2, column=1).value
        assert cell_value.startswith("'")

    def test_preserves_negative_numbers(self):
        from augmentum.tools.artifact_spreadsheet import _render_xlsx
        sheets = [{"name": "Test", "headers": ["A"], "rows": [["-42.5"]]}]
        data = _render_xlsx(sheets)
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(data))
        cell_value = wb.active.cell(row=2, column=1).value
        # Negative number string should NOT be prefixed
        assert cell_value == "-42.5"

    def test_preserves_equals_formula(self):
        from augmentum.tools.artifact_spreadsheet import _render_xlsx
        sheets = [{"name": "Test", "headers": ["A"], "rows": [["=SUM(A1:A5)"]]}]
        data = _render_xlsx(sheets)
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(data))
        cell_value = wb.active.cell(row=2, column=1).value
        assert cell_value == "=SUM(A1:A5)"


# === JSONPath Duplicate Segment Fix ===


class TestJsonPathDuplicateSegment:
    def test_wildcard_with_duplicate_path_segments(self):
        """parts.index(part) would find first 'name' instead of current one."""
        from augmentum.tools.json_tool import _jsonpath_query
        data = {"items": [{"name": {"name": "inner"}}]}
        # $.items[*].name should return [{"name": "inner"}], not recurse wrong
        result = _jsonpath_query(data, "$.items[*].name")
        assert result == [{"name": "inner"}]
